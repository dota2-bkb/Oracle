import streamlit as st
from database import get_db
from models import Match, Team, PlayerPerformance, Player, PickBan, League, PlayerAlias
from services.hero_manager import HeroManager
from services.patch_manager import PatchManager
from views.components import render_bp_visual, generate_bp_image, generate_bp_grid_image
from sqlalchemy import desc, func, or_
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- Helper Function for Statistics Sheet ---
def create_shared_stats_sheet(wb, matches, db, hm, team_name=None):
    ws_stats = wb.create_sheet("统计数据")
    
    # 3.1 Win Rates
    total = len(matches)
    if total > 0:
        wins = sum(1 for m in matches if m.win)
        
        rad_m = [m for m in matches if m.is_radiant]
        rad_wins = sum(1 for m in rad_m if m.win)
        
        dire_m = [m for m in matches if not m.is_radiant]
        dire_wins = sum(1 for m in dire_m if m.win)
        
        fp_m = [m for m in matches if m.first_pick]
        fp_wins = sum(1 for m in fp_m if m.win)
        
        sp_m = [m for m in matches if not m.first_pick]
        sp_wins = sum(1 for m in sp_m if m.win)
        
        stats_data = [
            ["统计项", "场次", "胜场", "胜率"],
            ["总计", total, wins, f"{wins/total:.1%}"],
            ["天辉", len(rad_m), rad_wins, f"{rad_wins/len(rad_m):.1%}" if rad_m else "0%"],
            ["夜魇", len(dire_m), dire_wins, f"{dire_wins/len(dire_m):.1%}" if dire_m else "0%"],
            ["先选", len(fp_m), fp_wins, f"{fp_wins/len(fp_m):.1%}" if fp_m else "0%"],
            ["后选", len(sp_m), sp_wins, f"{sp_wins/len(sp_m):.1%}" if sp_m else "0%"]
        ]
        
        for r_idx, row_data in enumerate(stats_data, start=1):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_stats.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1: cell.font = Font(bold=True)

    # 3.2 Top 10 Common Heroes & Combinations
    ws_stats.cell(row=8, column=1, value="常用英雄及最佳搭档 (Top 10)").font = Font(bold=True)
    # Update Header with Win Rate for partners
    ws_stats.append(["排名", "英雄", "出场次数", "胜率", "最佳搭档1", "场次", "胜率", "最佳搭档2", "场次", "胜率", "最佳搭档3", "场次", "胜率"])
    
    # Calc Logic
    pick_counts = {}
    hero_wins = {}
    hero_partners = {} # hid -> {partner_id: count}
    
    for m in matches:
        my_side = 0 if m.is_radiant else 1
        my_picks = [pb.hero_id for pb in m.pick_bans if pb.is_pick and pb.team_side == my_side]
        
        for hid in my_picks:
            pick_counts[hid] = pick_counts.get(hid, 0) + 1
            if m.win:
                hero_wins[hid] = hero_wins.get(hid, 0) + 1
                
            if hid not in hero_partners: hero_partners[hid] = {}
            for pid in my_picks:
                if hid != pid:
                    hero_partners[hid][pid] = hero_partners[hid].get(pid, 0) + 1
                    
    sorted_heroes = sorted(pick_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    
    start_row = 10
    for i, (hid, count) in enumerate(sorted_heroes, start=1):
        h_data = hm.get_hero(hid)
        h_name = h_data.get('slang') or h_data.get('cn_name')
        wr = hero_wins.get(hid, 0) / count
        
        row_vals = [i, h_name, count, f"{wr:.1%}"]
        
        # Partners
        partners = hero_partners.get(hid, {})
        # Calculate win rates for partners
        partner_stats = []
        for pid, p_count in partners.items():
            # Count wins where both hid and pid were on My Team
            wins_with_partner = 0
            for m in matches:
                my_side = 0 if m.is_radiant else 1
                my_picks = [pb.hero_id for pb in m.pick_bans if pb.is_pick and pb.team_side == my_side]
                if hid in my_picks and pid in my_picks:
                    if m.win:
                        wins_with_partner += 1
            
            p_wr = wins_with_partner / p_count if p_count > 0 else 0
            partner_stats.append((pid, p_count, p_wr))

        sorted_partners = sorted(partner_stats, key=lambda x: x[1], reverse=True)[:3]
        
        for pid, p_count, p_wr in sorted_partners:
            p_data = hm.get_hero(pid)
            p_name = p_data.get('slang') or p_data.get('cn_name')
            row_vals.extend([p_name, p_count, f"{p_wr:.1%}"])
            
        # Write Row
        for col, val in enumerate(row_vals, start=1):
            ws_stats.cell(row=start_row + i - 1, column=col, value=val)

    # 3.3 Signature Heroes by Position (1-5) - VERTICAL LAYOUT
    ws_stats.cell(row=start_row + 12, column=1, value="各位置绝活列表").font = Font(bold=True)
    
    # Identify Main Players
    # Logic:
    # 1. Try to find players manually assigned to this team & position in DB.
    # 2. If multiple manual players, use the one in the most recent match.
    # 3. If no manual player, fallback to most frequent player in matches.
    
    # Pre-calculate fallback counts
    pos_player_counts = {i: {} for i in range(1, 6)}
    for m in matches: # Use all matches passed in
        my_side = 0 if m.is_radiant else 1
        my_ps = [p for p in m.players if p.team_side == my_side]
        for p in my_ps:
            if p.position and 1 <= p.position <= 5 and p.account_id:
                pos_player_counts[p.position][p.account_id] = pos_player_counts[p.position].get(p.account_id, 0) + 1
    
    # Prepare Manual Players Map
    # Resolve Team ID from team_name
    manual_players = {i: [] for i in range(1, 6)}
    if team_name:
        target_team = db.query(Team).filter(Team.name == team_name).first()
        if target_team:
            # Get players in this team with default_pos set
            team_players = db.query(Player).filter(Player.team_id == target_team.team_id, Player.default_pos != None).all()
            for p in team_players:
                if 1 <= p.default_pos <= 5:
                    manual_players[p.default_pos].append(p.account_id)
    
    main_players = {}
    
    for pos in range(1, 6):
        candidates = manual_players[pos]
        
        if candidates:
            # Case 1: Manual assignment exists
            if len(candidates) == 1:
                main_players[pos] = candidates[0]
            else:
                # Conflict: Multiple players for this pos. Find most recent.
                # Sort candidates by "last match time" in the provided matches list
                best_candidate = None
                best_time = None
                
                # We need to scan matches to find latest appearance
                # Matches are sorted by time (asc) in 'matches_asc' or we iterate 'matches' (which might be desc?)
                # 'matches' passed to this function is usually desc (latest first) based on call site, but let's be safe.
                # Actually, let's just look at 'matches' order.
                
                found = False
                # Sort matches desc by time just to be sure
                matches_desc = sorted(matches, key=lambda m: m.match_time, reverse=True)
                
                for m in matches_desc:
                    if found: break
                    my_side = 0 if m.is_radiant else 1
                    my_ps = [p for p in m.players if p.team_side == my_side]
                    
                    for p in my_ps:
                        if p.account_id in candidates:
                            # Found the most recent one
                            main_players[pos] = p.account_id
                            found = True
                            break
                
                if not found:
                    # None of them played in these matches? Just pick the first one from DB
                    main_players[pos] = candidates[0]
                    
        else:
            # Case 2: No manual assignment, use statistics (fallback)
            counts = pos_player_counts[pos]
            if counts:
                main_players[pos] = max(counts, key=counts.get)
            
    base_r = start_row + 13
    # Vertical Layout: 5 Columns (one per position)
    
    for pos in range(1, 6):
        # Calculate column offset for each position (3 columns wide per position + 1 gap)
        col_offset = 1 + (pos - 1) * 4 
        
        r = base_r
        ws_stats.cell(row=r, column=col_offset, value=f"{pos}号位").font = Font(bold=True)
        r += 1
        
        acc_id = main_players.get(pos)
        
        if not acc_id:
            ws_stats.cell(row=r, column=col_offset, value="未识别")
            continue
            
        # Get Name
        alias = db.query(PlayerAlias).filter(PlayerAlias.account_id == acc_id).first()
        p_info = db.query(Player).filter(Player.account_id == acc_id).first()
        p_name = "未知"
        if alias and alias.player: p_name = alias.player.name
        elif p_info: p_name = p_info.name
        else: p_name = str(acc_id)
        
        ws_stats.cell(row=r, column=col_offset, value=p_name).font = Font(bold=True, color="0000FF")
        r += 1
        
        # Headers for this player
        ws_stats.cell(row=r, column=col_offset, value="英雄")
        ws_stats.cell(row=r, column=col_offset+1, value="场次")
        ws_stats.cell(row=r, column=col_offset+2, value="胜率")
        r += 1
        
        # Get Top Heroes for this player in these matches
        p_heroes = {} # hid -> {picks, wins}
        for m in matches:
             p_rec = next((p for p in m.players if p.account_id == acc_id), None)
             if p_rec:
                 hid = p_rec.hero_id
                 if hid not in p_heroes: p_heroes[hid] = {'picks': 0, 'wins': 0}
                 p_heroes[hid]['picks'] += 1
                 
                 # Check win
                 radiant_won = (m.is_radiant == m.win)
                 player_won = (p_rec.team_side == 0 and radiant_won) or (p_rec.team_side == 1 and not radiant_won)
                 if player_won:
                     p_heroes[hid]['wins'] += 1

        # 原来只截取前 5 个英雄，现在改为：有多少就展示多少
        sorted_ph = sorted(p_heroes.items(), key=lambda x: x[1]['picks'], reverse=True)
        
        for hid, stats in sorted_ph:
            h_data = hm.get_hero(hid)
            h_name = h_data.get('slang') or h_data.get('cn_name')
            cnt = stats['picks']
            wr = stats['wins'] / cnt if cnt > 0 else 0
            
            ws_stats.cell(row=r, column=col_offset, value=h_name)
            ws_stats.cell(row=r, column=col_offset+1, value=cnt)
            ws_stats.cell(row=r, column=col_offset+2, value=f"{wr:.1%}")
            r += 1

def generate_detailed_excel_export(matches, team_name, db, hm):
    """
    Template 1: Detailed Match & Stats
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Sort matches by time ascending (Old -> New) for the horizontal layout
    # User said: "left to right sequentially increasing time"
    matches_asc = sorted(matches, key=lambda m: m.match_time)
    
    # --- Helper to create Match Sheet ---
    def create_match_sheet(sheet_name, filter_func):
        ws = wb.create_sheet(sheet_name)
        filtered_matches = [m for m in matches_asc if filter_func(m)]
        
        if not filtered_matches:
            ws.cell(row=1, column=1, value="无符合条件的比赛数据")
            return

        # Set Row Headers
        headers = ["比赛ID", "日期", "联赛", "天辉队伍", "夜魇队伍", "获胜方", "BP详情"]
        for r, header in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=1, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        # Set Row Heights
        ws.row_dimensions[7].height = 200 # Approx height for BP image (adjust as needed)
        ws.column_dimensions['A'].width = 15

        for idx, m in enumerate(filtered_matches):
            col_idx = idx + 2 # Start from Column B
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 50 # Width for BP image

            # 1. Match ID
            ws.cell(row=1, column=col_idx, value=str(m.match_id)).alignment = Alignment(horizontal='center')
            
            # 2. Date
            ws.cell(row=2, column=col_idx, value=m.match_time.strftime('%Y-%m-%d')).alignment = Alignment(horizontal='center')
            
            # 3. League (Lazy Load name)
            l_name = "未知联赛"
            if m.league_id:
                lg = db.query(League).filter(League.league_id == m.league_id).first()
                if lg: l_name = lg.name
            ws.cell(row=3, column=col_idx, value=l_name).alignment = Alignment(horizontal='center', wrap_text=True)

            # 4. Radiant
            r_name = m.team_name if m.is_radiant else m.opponent_name
            ws.cell(row=4, column=col_idx, value=r_name).alignment = Alignment(horizontal='center')
            
            # 5. Dire
            d_name = m.opponent_name if m.is_radiant else m.team_name
            ws.cell(row=5, column=col_idx, value=d_name).alignment = Alignment(horizontal='center')
            
            # 6. Winner (With crown)
            winner_name = m.team_name if m.win else m.opponent_name
            ws.cell(row=6, column=col_idx, value=f"👑 {winner_name}").alignment = Alignment(horizontal='center')
            
            # 7. BP Image
            # Determine params for generation
            rad_name = m.team_name if m.is_radiant else m.opponent_name
            dire_name = m.opponent_name if m.is_radiant else m.team_name
            is_radiant_first = (m.is_radiant == m.first_pick)
            
            # Add winner mark to team name passed to image gen?
            # User requirement: "BP图片中谁获胜了需要文字中添加信息"
            # We modify the names passed to the image generator
            rad_disp = f"👑 {rad_name}" if (m.is_radiant == m.win) else rad_name
            dire_disp = f"👑 {dire_name}" if (m.is_radiant != m.win) else dire_name
            
            try:
                img_pil = generate_bp_image(m.pick_bans, rad_disp, dire_disp, hm, first_pick_radiant=is_radiant_first)
                if img_pil:
                    # Scaling
                    w_target = 300
                    h_target = int(w_target * (img_pil.height / img_pil.width))
                    
                    img_resized = img_pil.resize((w_target, h_target))
                    
                    # Convert to Bytes for OpenPyXL
                    img_byte_arr = BytesIO()
                    img_resized.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    xl_img = XLImage(img_byte_arr)
                    
                    # Anchor to cell
                    ws.add_image(xl_img, f"{col_letter}7")
                    
                    ws.row_dimensions[7].height = h_target * 0.75
            except Exception as e:
                print(f"Error generating image for excel: {e}")
                ws.cell(row=7, column=col_idx, value="图片生成失败")

    # Sheet 1: Analyzed Team First Pick
    create_match_sheet(f"{team_name}-先选", lambda m: m.first_pick)
    
    # Sheet 2: Analyzed Team Second Pick
    create_match_sheet(f"{team_name}-后选", lambda m: not m.first_pick)
    
    # --- Sheet 3: 统计信息 (Stats) ---
    # 这里必须传入 team_name，才能在“各位置绝活列表”中优先使用 Player Manager 中手动配置的主力位置
    create_shared_stats_sheet(wb, matches, db, hm, team_name=team_name)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_template_2(matches, team_name, db, hm):
    """
    Template 2: Grid Style BP Image (Vertical List, No Text, Original Width)
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Sort matches by time ascending
    matches_asc = sorted(matches, key=lambda m: m.match_time)
    
    # --- Helper to create Match Sheet ---
    def create_match_sheet(sheet_name, filter_func):
        ws = wb.create_sheet(sheet_name)
        filtered_matches = [m for m in matches_asc if filter_func(m)]
        
        # Header
        ws.cell(row=1, column=1, value=sheet_name).font = Font(bold=True, size=14)
        
        if not filtered_matches:
            ws.cell(row=2, column=1, value="无符合条件的比赛数据")
            return

        # Set Column A Width to accommodate image width
        # Original width ~1280px. 50% = 640px. Excel width ~90
        ws.column_dimensions['A'].width = 90

        for idx, m in enumerate(filtered_matches):
            row_idx = idx + 2 # Start from row 2 (row 1 is header)
            
            # --- Generate Grid Image ---
            # Analyzed Team on LEFT
            left_is_radiant = m.is_radiant
            left_team_name = m.team_name
            right_team_name = m.opponent_name
            
            # Winner logic
            winner_is_left = m.win 
            
            # First Pick logic
            left_is_first_pick = m.first_pick
            
            try:
                img_pil = generate_bp_grid_image(
                    m.pick_bans, 
                    left_team_name, right_team_name, 
                    hm, 
                    left_is_radiant=left_is_radiant,
                    left_is_first_pick=left_is_first_pick,
                    winner_is_left=winner_is_left
                )
                
                if img_pil:
                    # Resize to 50%
                    new_w = int(img_pil.width * 0.5)
                    new_h = int(img_pil.height * 0.5)
                    img_resized = img_pil.resize((new_w, new_h))
                    
                    img_byte_arr = BytesIO()
                    img_resized.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    xl_img = XLImage(img_byte_arr)
                    ws.add_image(xl_img, f"A{row_idx}")
                    
                    # Set Row Height
                    ws.row_dimensions[row_idx].height = new_h * 0.75
                    
            except Exception as e:
                print(f"Error generating grid image: {e}")
                ws.cell(row=row_idx, column=1, value="图片生成失败")

    # Sheet 1: Analyzed Team First Pick
    create_match_sheet(f"{team_name}-先选", lambda m: m.first_pick)
    
    # Sheet 2: Analyzed Team Second Pick
    create_match_sheet(f"{team_name}-后选", lambda m: not m.first_pick)
    
    # --- Sheet 3: 统计信息 (Stats) ---
    create_shared_stats_sheet(wb, matches, db, hm, team_name=team_name)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_template_3(matches, team_name, db, hm):
    """
    Template 3: Pure Text Log (Detailed BP & Positions)
    """
    wb = Workbook()
    # Default sheet
    ws = wb.active
    ws.title = "战绩详情"
    
    # 1. Styles
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_green  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid") # Light Green
    fill_blue   = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Light Blue (Win)
    fill_red    = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid") # Pink (Loss)
    
    font_header = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 2. Headers
    # Groups: [Meta 4] [My Team 5] [Opponent 5] [Side] [Bans]
    # [比赛编号, 时间, 对手, 胜负, 我方pick1-5, 对方pick1-5, 阵营, 对方ban1-7, 我方ban1-7]
    
    # Row 1: High Level Headers
    ws.merge_cells('A1:D1') # Meta
    ws.cell(row=1, column=1, value="比赛信息").alignment = alignment_center
    
    ws.merge_cells('E1:I1') # My Team
    ws.cell(row=1, column=5, value=team_name).alignment = alignment_center
    ws.cell(row=1, column=5).font = Font(bold=True, size=12)
    
    ws.merge_cells('J1:O1') # Opponent
    ws.cell(row=1, column=10, value="对手").alignment = alignment_center
    
    # FP Bans: 7 columns (P to V)
    ws.merge_cells('P1:V1') 
    ws.cell(row=1, column=16, value="1选Ban").alignment = alignment_center
    ws.cell(row=1, column=16).fill = fill_yellow
    
    # SP Bans: 7 columns (W to AC)
    ws.merge_cells('W1:AC1') 
    ws.cell(row=1, column=23, value="2选Ban").alignment = alignment_center
    ws.cell(row=1, column=23).fill = fill_green 

    # Row 2: Detail Headers
    headers = [
        "编号", "时间", "对手", "胜负",              # Meta
        "我方1", "我方2", "我方3", "我方4", "我方5",
        "阵营",     # My Team Picks
        "对手1", "对手2", "对手3", "对手4", "对手5",    # Opponent Picks                       
        "对手Ban1", "对手Ban2", "对手Ban3", "对手Ban4", "对手Ban5", "对手Ban6", "对手Ban7",
        "我方Ban1", "我方Ban2", "我方Ban3", "我方Ban4", "我方Ban5", "我方Ban6", "我方Ban7"
    ]
    
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = font_header
        c.alignment = alignment_center
        c.border = border_thin

    # 3. Data Rows
    # Using 'matches' which is typically passed as Descending time (Latest first).
    
    # Global Order Constants for Coloring
    # DOTA 2 (Current Captains Mode - or user approximation)
    # Pick Orders (1-based Global):
    # FP Picks: 9 (1st), 13, 16, 17, 24
    # SP Picks: 8 (1st), 14, 15, 18, 23
    #
    # User Logic:
    # "1选方1选" -> FP Team's 1st Pick (Order 9) -> Yellow
    # "对方counter 1 2选" -> Opponent's 1st & 2nd Pick
    #   If Opponent is SP: Orders 8, 14 -> Green
    #   If Opponent is FP (My Team is SP): Orders 9, 13 -> Green
    
    # Correction per user request:
    # "1 选人高亮没有做：每场比赛 第8手黄色格子 第9 13手绿色格子"
    # Wait, if User says Order 8 is Yellow, then Order 8 must be considered "1选" (First Pick of Phase).
    # In our system, Order 8 is SP's first pick. Order 9 is FP's first pick.
    # User might be using a system where Order 8 is the FIRST pick of the draft?
    # Standard Dota 2: First Pick Team picks at 8? No, at 5, 8?
    # Let's strictly follow User's Explicit Instruction:
    # "第8手黄色格子" -> Order 8 Yellow.
    # "第9 13手绿色格子" -> Order 9, 13 Green.
    
    COLOR_YELLOW_PICK_ORDER = 8
    COLOR_GREEN_PICK_ORDERS = [9, 13]
    
    # Ban Colors:
    # "1 4 7 手黄色 2 3 5 6 手绿色"
    COLOR_YELLOW_BAN_ORDERS = [1, 4, 7]
    COLOR_GREEN_BAN_ORDERS = [2, 3, 5, 6]

    row_idx = 3
    for idx, m in enumerate(matches, start=1):
        # A: Index
        ws.cell(row=row_idx, column=1, value=idx).border = border_thin
        
        # B: Time
        ws.cell(row=row_idx, column=2, value=m.match_time.strftime('%m-%d')).border = border_thin
        
        # C: Opponent
        ws.cell(row=row_idx, column=3, value=m.opponent_name).border = border_thin
        
        # D: Result
        res_str = "胜" if m.win else "负"
        res_cell = ws.cell(row=row_idx, column=4, value=res_str)
        res_cell.fill = fill_blue if m.win else fill_red
        res_cell.border = border_thin
        res_cell.alignment = alignment_center
        
        # --- Process Players for Positions 1-5 ---
        def get_pos_map(is_radiant):
            # Returns {pos (1-5): hero_id}
            pmap = {}
            side = 0 if is_radiant else 1
            
            # Filter players on this side
            # Note: m.players is a list of PlayerPerformance objects
            team_players = [p for p in m.players if p.team_side == side]
            
            for p in team_players:
                # Use 'position' field (1-5)
                # Fallback to internal logic if 0/None?
                # Assuming data is clean enough or we accept gaps
                if p.position and 1 <= p.position <= 5:
                    pmap[p.position] = p.hero_id
            return pmap

        my_is_radiant = m.is_radiant
        my_pmap = get_pos_map(my_is_radiant)
        opp_pmap = get_pos_map(not my_is_radiant)
        
        # Logic for Highlighting
        # Map hero_id -> Pick Order to determine if it needs color
        # PickBan list has this info
        pick_order_map = {} # hero_id -> global_order
        ban_order_map = {}
        for pb in m.pick_bans:
            if pb.is_pick:
                pick_order_map[pb.hero_id] = pb.order + 1 # 1-based
            else:
                ban_order_map[pb.hero_id] = pb.order + 1
        
        # Determine who is FP
        # m.first_pick is True if 'team_name' (My Team) was First Pick
        i_am_fp = m.first_pick
        
        # E-I: My Team Pos 1-5
        for p in range(1, 6):
            col = 4 + p # E is 5
            hid = my_pmap.get(p)
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border_thin
            
            if hid:
                h_data = hm.get_hero(hid)
                h_name = h_data.get('slang') or h_data.get('cn_name')
                cell.value = h_name
                
                # Color Logic
                order = pick_order_map.get(hid)
                if order:
                    if order == COLOR_YELLOW_PICK_ORDER:
                        cell.fill = fill_yellow
                    elif order in COLOR_GREEN_PICK_ORDERS:
                        cell.fill = fill_green

        # K-O: Opponent Pos 1-5  (列 10-14)
        for p in range(1, 6):
            col = 10 + p  # 10..14
            hid = opp_pmap.get(p)
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border_thin
            
            if hid:
                h_data = hm.get_hero(hid)
                h_name = h_data.get('slang') or h_data.get('cn_name')
                cell.value = h_name
                
                # Color Logic for Opponent
                order = pick_order_map.get(hid)
                if order:
                    if order == COLOR_YELLOW_PICK_ORDER:
                        cell.fill = fill_yellow
                    elif order in COLOR_GREEN_PICK_ORDERS:
                        cell.fill = fill_green

        # 阵营列（My Side），放在第 15 列
        side_str = "天辉" if my_is_radiant else "夜魇"
        side_cell = ws.cell(row=row_idx, column=10, value=side_str)
        # 天辉=绿色粗体，夜魇=红色粗体
        if side_str == "天辉":
            side_cell.font = Font(color="008000", bold=True)
        else:
            side_cell.font = Font(color="FF0000", bold=True)
        side_cell.border = border_thin
        side_cell.alignment = alignment_center

        # --- Bans ---
        # 仍按全局顺位：1/4/7 黄，2/3/5/6 绿
        my_side_int = 0 if my_is_radiant else 1
        my_bans = []
        opp_bans = []
        
        sorted_pbs = sorted(m.pick_bans, key=lambda x: x.order)
        
        for pb in sorted_pbs:
            if not pb.is_pick:
                h_data = hm.get_hero(pb.hero_id)
                h_name = h_data.get('slang') or h_data.get('cn_name')
                order = pb.order + 1
                item = {'name': h_name, 'order': order}
                
                if pb.team_side == my_side_int:
                    my_bans.append(item)
                else:
                    opp_bans.append(item)
                    
        # 对方 Ban：P-V 列（16-22）
        for i in range(7):
            col = 16 + i
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border_thin
            
            if i < len(opp_bans):
                b = opp_bans[i]
                cell.value = b['name']
                o = b['order']
                if o in COLOR_YELLOW_BAN_ORDERS: cell.fill = fill_yellow
                elif o in COLOR_GREEN_BAN_ORDERS: cell.fill = fill_green
            else:
                cell.value = "-"
            
        # 我方 Ban：W-AC 列（23-29）
        for i in range(7):
            col = 23 + i
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border_thin
            
            if i < len(my_bans):
                b = my_bans[i]
                cell.value = b['name']
                o = b['order']
                if o in COLOR_YELLOW_BAN_ORDERS: cell.fill = fill_yellow
                elif o in COLOR_GREEN_BAN_ORDERS: cell.fill = fill_green
            else:
                cell.value = "-"

        row_idx += 1

    # Auto-adjust widths (rough)
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    for c in range(5, 30):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # --- Sheet 2: 统计信息 (Stats) ---
    create_shared_stats_sheet(wb, matches, db, hm)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_template_4(matches, team_name, db, hm):
    """
    Template 4: Review Template (Text Template 2)
    Customized per user request:
    - Headers are dynamic player names for My Team
    - Content: HeroSlang + #Order
    - Ban: 7 bans. First Ban side is Yellow.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "复盘详情"
    
    # Styles
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_green  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid") # Light Green (Team)
    fill_blue   = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") # Win
    fill_red    = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid") # Loss
    fill_gray_red = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid") # Dire text like
    fill_gray_green = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid") # Radiant text like
    
    font_header = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Identify Main Players for Header ---
    # Reuse logic from analysis tab roughly
    pos_player_counts = {i: {} for i in range(1, 6)}
    for m in matches:
        my_side = 0 if m.is_radiant else 1
        my_ps = [p for p in m.players if p.team_side == my_side]
        for p in my_ps:
            if p.position and 1 <= p.position <= 5 and p.account_id:
                pos_player_counts[p.position][p.account_id] = pos_player_counts[p.position].get(p.account_id, 0) + 1
    
    # Manual DB Lookup
    target_team = db.query(Team).filter(Team.name == team_name).first()
    manual_players = {i: [] for i in range(1, 6)}
    if target_team:
        team_players = db.query(Player).filter(Player.team_id == target_team.team_id, Player.default_pos != None).all()
        for p in team_players:
            if 1 <= p.default_pos <= 5:
                manual_players[p.default_pos].append(p.account_id)
    
    main_players_map = {} # pos -> name
    matches_desc = sorted(matches, key=lambda m: m.match_time, reverse=True)
    
    for pos in range(1, 6):
        final_acc_id = None
        candidates = manual_players[pos]
        if candidates:
            if len(candidates) == 1:
                final_acc_id = candidates[0]
            else:
                found = False
                for m in matches_desc:
                    if found: break
                    my_side = 0 if m.is_radiant else 1
                    my_ps = [p for p in m.players if p.team_side == my_side]
                    for p in my_ps:
                        if p.account_id in candidates:
                            final_acc_id = p.account_id
                            found = True
                            break
                if not found: final_acc_id = candidates[0]
        else:
            counts = pos_player_counts[pos]
            if counts:
                final_acc_id = max(counts, key=counts.get)
        
        # Get Name
        p_name = f"{pos}号位"
        if final_acc_id:
            alias = db.query(PlayerAlias).filter(PlayerAlias.account_id == final_acc_id).first()
            p_info = db.query(Player).filter(Player.account_id == final_acc_id).first()
            if alias and alias.player: p_name = alias.player.name
            elif p_info: p_name = p_info.name
            else: p_name = str(final_acc_id)
        
        main_players_map[pos] = p_name

    # --- Headers ---
    # Row 1: Team Name Title
    ws.merge_cells('A1:AC1') 
    ws.cell(row=1, column=1, value=f"{team_name} 数据复盘").alignment = alignment_center
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)

    # Row 2: Detail Headers
    # A-D: Base
    # E-I: My Team (Player Names)
    # J: Side
    # K-O: Opponent (Pos 1-5)
    # P-V: Opp Ban
    # W-AC: My Ban
    
    headers_r2 = ["比赛编号", "时间", "对手", "胜负"]
    for i in range(1, 6): headers_r2.append(main_players_map.get(i, f"{i}号位"))
    headers_r2.append("阵营")
    # K-O: Opponent (Pos 1-5) -> Change to: 1 - PlayerID, 2 - PlayerID ...
    # We can't know dynamic opponent names easily for ALL matches in one header row.
    # User requirement: "目标队伍12345 要写 1 - 选手ID"
    # Wait, "目标队伍" usually means the Analyzed Team (My Team).
    # My Team headers are already dynamic names (E-I).
    # Maybe user means Opponent headers should be "1号位 - ID"? No, ID changes every match.
    # Re-reading: "b. 排列上... 而是按照 |目标队伍PICK 12345|对手PICK12345|对手ban|目标队伍ban|"
    # AND "并且目标队伍12345 要写 1 - 选手ID"
    # This refers to the HEADER row for My Team columns?
    # Yes, currently I put "Name". User wants "1 - Name", "2 - Name"?
    # "1 - 选手ID". If I have Name, use Name.
    
    my_team_headers = []
    for i in range(1, 6):
        p_name = main_players_map.get(i, "Unknown")
        my_team_headers.append(f"{i} - {p_name}")

    # Re-arrange columns as requested:
    # | My Team Pick (5) | Opp Pick (5) | Opp Ban (7) | My Ban (7) |
    # Wait, original was: | My | Side | Opp | ...
    # New request: | My | Opp | Opp Ban | My Ban |
    # "阵营" column is missing in the new description?
    # "a. 阵营页面 天辉为绿色粗体，夜魇为红色粗体" -> This implies "阵营" column still exists?
    # Or maybe "阵营页面" means "阵营" column cell content?
    # Let's keep "阵营" column, probably between My and Opp or at end?
    # User list: |目标队伍PICK|对手PICK|对手ban|目标队伍ban|
    # It skipped "Side", "MatchID", "Time", "Result".
    # I assume A-D (Base Info) stays.
    # Where does "Side" go? Maybe remove it and use color coding on Team Name?
    # But user said "阵营页面 天辉为绿色..." which strongly implies a specific column or page.
    # Let's keep "Side" column (J) but move it? Or keep it in middle?
    # The user request "排列上 总体上不是... 而是..." lists the Pick/Ban blocks.
    # Block 1: My Pick
    # Block 2: Opp Pick
    # Block 3: Opp Ban
    # Block 4: My Ban
    # Previous was: My Pick | Side | Opp Pick | Opp Ban | My Ban
    # So "Side" is the main difference. Let's move Side to before My Pick? Or keep it at J?
    # If I follow strictly: "My | Opp | Opp Ban | My Ban" -> Side is gone.
    # But "a. 阵营页面..." request implies it exists.
    # I will put Side column J between My and Opp, as it separates them nicely.
    
    # Revised Column Structure:
    # A-D: Base
    # E-I: My Team (1-5)
    # J: Side
    # K-O: Opponent (1-5)
    # P-V: Opp Ban (7)
    # W-AC: My Ban (7)
    
    # This matches the user's "My | Opp | Opp Ban | My Ban" block order (ignoring Side/Base for a moment).
    # Previous implementation was: My | Side | Opp | Opp Ban | My Ban.
    # Wait, previous was: My | Side | Opp | Opp Ban | My Ban.
    # User says: |目标队伍PICK|对手PICK|对手ban|目标队伍ban|
    # It matches my previous implementation except maybe the internal order of bans?
    # Ah, "1选ban|2选ban" vs "对手ban|目标队伍ban".
    # Previous code:
    # headers_r2.append("对手Ban人") ...
    # headers_r2.append("己方Ban人") ...
    # It seems I already did "Opp Ban" then "My Ban".
    # Maybe user thinks I did "1st Ban" "2nd Ban"?
    # Template 3 was "1选方Ban" "2选方Ban".
    # Template 4 implementation (current): "对手Ban人" "己方Ban人".
    # So the order is already correct?
    # Let's check "b. 排列上...". Maybe they want to remove "Side" column?
    # I'll stick to: A-D Base, E-I My, J Side, K-O Opp, P-V OppBan, W-AC MyBan.
    
    # Update Header content for My Team
    headers_r2 = ["比赛编号", "时间", "对手", "胜负"]
    headers_r2.extend(my_team_headers)
    headers_r2.append("阵营")
    for i in range(1, 6): headers_r2.append(f"{i}号位")
    
    headers_r2.append("对手Ban人")
    for _ in range(6): headers_r2.append("")
    
    headers_r2.append("己方Ban人")
    for _ in range(6): headers_r2.append("")
    
    for i, h in enumerate(headers_r2, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = font_header
        c.alignment = alignment_center
        c.border = border_thin
    
    # Merge Ban Headers
    ws.merge_cells('P2:V2')
    ws.merge_cells('W2:AC2')
    
    # --- Data Rows ---
    row_idx = 3
    for idx, m in enumerate(matches, start=1):
        # A: Index
        ws.cell(row=row_idx, column=1, value=m.match_id).border = border_thin
        
        # B: Time
        ws.cell(row=row_idx, column=2, value=m.match_time.strftime('%m.%d')).border = border_thin
        
        # C: Opponent
        ws.cell(row=row_idx, column=3, value=m.opponent_name).border = border_thin
        
        # D: Result
        res_str = "胜" if m.win else "负"
        res_cell = ws.cell(row=row_idx, column=4, value=res_str)
        res_cell.fill = fill_green if m.win else fill_red
        res_cell.border = border_thin
        res_cell.alignment = alignment_center
        
        # Helper for content
        def get_hero_cell_val(hid, pick_order_map):
             if not hid: return "-"
             h_data = hm.get_hero(hid)
             h_name = h_data.get('slang') or h_data.get('cn_name')
             # Append Order #N
             order = pick_order_map.get(hid, "?")
             return f"{h_name}{order}"

        pick_order_map = {pb.hero_id: pb.order + 1 for pb in m.pick_bans if pb.is_pick}
        
        # My Side & Opp Side
        my_is_radiant = m.is_radiant
        
        # E-I: My Team
        # Find hero for each position 1-5
        # m.players has position info.
        my_side_int = 0 if my_is_radiant else 1
        my_pmap = {}
        for p in m.players:
            if p.team_side == my_side_int and p.position and 1 <= p.position <= 5:
                my_pmap[p.position] = p.hero_id
                
        for i in range(1, 6):
            hid = my_pmap.get(i)
            val = get_hero_cell_val(hid, pick_order_map)
            ws.cell(row=row_idx, column=4+i, value=val).border = border_thin
            
        # J: Side
        # Requirement: "天辉为绿色粗体，夜魇为红色粗体"
        side_str = "天辉" if my_is_radiant else "夜魇"
        side_cell = ws.cell(row=row_idx, column=10, value=side_str)
        if my_is_radiant:
             side_cell.font = Font(color="008000", bold=True) # Green
        else:
             side_cell.font = Font(color="FF0000", bold=True) # Red
        side_cell.border = border_thin
        side_cell.alignment = alignment_center
        
        # K-O: Opponent
        opp_side_int = 1 if my_is_radiant else 0
        opp_pmap = {}
        for p in m.players:
             if p.team_side == opp_side_int and p.position and 1 <= p.position <= 5:
                opp_pmap[p.position] = p.hero_id
        
        for i in range(1, 6):
            hid = opp_pmap.get(i)
            val = get_hero_cell_val(hid, pick_order_map)
            ws.cell(row=row_idx, column=10+i, value=val).border = border_thin

        # --- Bans ---
        i_am_first_ban = m.first_pick 
        opp_is_first_ban = not i_am_first_ban
        
        # Collect bans by team
        my_bans = []
        opp_bans = []
        
        sorted_bans = sorted([pb for pb in m.pick_bans if not pb.is_pick], key=lambda x: x.order)
        
        for b in sorted_bans:
            # Check team side
            is_my_ban = (b.team_side == my_side_int)
            h_data = hm.get_hero(b.hero_id)
            h_name = h_data.get('slang') or h_data.get('cn_name')
            val = f"{h_name}{b.order + 1}"
            
            if is_my_ban: my_bans.append(val)
            else: opp_bans.append(val)
            
        # Fill P-V: Opponent Bans
        for i in range(7):
            col = 16 + i
            val = opp_bans[i] if i < len(opp_bans) else "-"
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border_thin
            if opp_is_first_ban: cell.fill = fill_yellow
            
        # Fill W-AC: My Bans
        for i in range(7):
            col = 23 + i
            val = my_bans[i] if i < len(my_bans) else "-"
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border_thin
            if i_am_first_ban: cell.fill = fill_yellow
            
        row_idx += 1

    # --- Auto Fit Column Width (Approximation) ---
    # Loop through all columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        
        # Sample first 20 rows to save time if large
        for cell in col[:20]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.5
        if adjusted_width > 50: adjusted_width = 50 # Cap
        ws.column_dimensions[column].width = adjusted_width

    # --- Sheet 2: Stats ---
    create_shared_stats_sheet(wb, matches, db, hm, team_name=team_name)
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def show():
    st.title("统计分析")
    
    db = next(get_db())
    hm = HeroManager()
    pm = PatchManager()
    
    # --- Sidebar Filters ---
    st.sidebar.header("分析配置")
    
    # Team Selection
    teams = [r[0] for r in db.query(Match.team_name).distinct().all()]
    selected_team = st.sidebar.selectbox("目标战队", options=teams)
    
    if not selected_team:
        st.info("请先选择战队。")
        db.close()
        return
    
    # League Selection (Updated Logic: Dynamic Filter based on Selected Team)
    # Filter matches for the selected team first to get relevant leagues
    team_matches_query = db.query(Match.league_id).filter(Match.team_name == selected_team).distinct()
    team_league_ids = [r[0] for r in team_matches_query.all()]
    
    # Fetch League details only for these IDs
    if team_league_ids:
        relevant_leagues = db.query(League).filter(League.league_id.in_(team_league_ids)).order_by(League.league_id.desc()).all()
        league_opts = {f"{l.name}": l.league_id for l in relevant_leagues}
    else:
        league_opts = {}
        
    selected_league_names = st.sidebar.multiselect("筛选联赛", options=list(league_opts.keys()))
    selected_league_ids = [league_opts[n] for n in selected_league_names] if selected_league_names else []

    # Date/Patch Filter
    patches = pm.get_all_patches()
    filter_mode = st.sidebar.radio("时间范围模式", ["按版本", "按日期"])
    
    start_date = None
    if filter_mode == "按版本":
        selected_patch = st.sidebar.selectbox("选择版本", patches)
        if selected_patch:
            start_date = pm.get_patch_date(selected_patch)
            st.sidebar.caption(f"起始日期: {start_date}")
    else:
        start_date = st.sidebar.date_input("起始日期", value=datetime.today().date() - timedelta(days=90))
        
    # Build Query
    query = db.query(Match).filter(Match.team_name == selected_team)
    
    if start_date:
        query = query.filter(Match.match_time >= start_date)
    
    if selected_league_ids:
        query = query.filter(Match.league_id.in_(selected_league_ids))
        
    matches = query.order_by(Match.match_time.desc()).all()
    
    if not matches:
        st.warning("该范围内无比赛数据。")
        db.close()
        return
    
    st.sidebar.success(f"已加载 {len(matches)} 场比赛")
    
    # --- Excel Export Button ---
    st.sidebar.markdown("---")
    st.sidebar.subheader("Excel 报告")
    
    # Template Selection
    export_template = st.sidebar.selectbox(
        "选择导出模版",
        options=["默认模版", "图片模板", "文字模板", "文字模板2"]
    )

    # Export Limit
    export_limit = st.sidebar.number_input(
        "导出条目数量 (最近 N 场)",
        min_value=1,
        max_value=len(matches),
        value=len(matches)
    )
    
    if st.sidebar.button("生成 Excel 报告"):
        with st.spinner("正在生成 Excel 报告..."):
            # Ensure filtering context is passed/used implicitly by passing 'matches' which is already filtered.
            excel_data = None
            
            # Slice matches for export only
            matches_to_export = matches[:export_limit]

            if "默认模版" in export_template:
                excel_data = generate_detailed_excel_export(matches_to_export, selected_team, db, hm)
            elif "图片模板" in export_template:
                excel_data = generate_template_2(matches_to_export, selected_team, db, hm)
            elif "文字模板2" in export_template:
                excel_data = generate_template_4(matches_to_export, selected_team, db, hm)
            elif "文字模板" in export_template:
                # Keep this last as it matches partially
                excel_data = generate_template_3(matches_to_export, selected_team, db, hm)
            
            if excel_data:
                st.sidebar.download_button(
                    label="📥 下载 Excel",
                    data=excel_data,
                    file_name=f"{selected_team}_{export_template.split(':')[0]}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # =================================================================
    # TABS
    # =================================================================
    tab_team, tab_player, tab_bp = st.tabs(["🛡️ 战队概况", "👤 选手绝活", "⛓️ BP 链条"])
    
    # -----------------------------------------------------------------
    # TAB 1: 战队概况
    # -----------------------------------------------------------------
    with tab_team:
        total = len(matches)
        wins = sum(1 for m in matches if m.win)
        
        rad_m = [m for m in matches if m.is_radiant]
        rad_wins = sum(1 for m in rad_m if m.win)
        
        dire_m = [m for m in matches if not m.is_radiant]
        dire_wins = sum(1 for m in dire_m if m.win)
        
        rad_wr = (sum(1 for m in rad_m if m.win) / len(rad_m) * 100) if rad_m else 0
        dire_wr = (sum(1 for m in dire_m if m.win) / len(dire_m) * 100) if dire_m else 0
        
        st.subheader("胜率统计")
        c1, c2, c3 = st.columns(3)
        c1.metric("总胜率", f"{(wins/total*100):.1f}%", f"{wins}胜 - {total-wins}负")
        c2.metric("天辉胜率", f"{rad_wr:.1f}%", f"{len(rad_m)}场")
        c3.metric("夜魇胜率", f"{dire_wr:.1f}%", f"{len(dire_m)}场")
        
        st.divider()
        
        # --- Hero Stats ---
        pick_counts = {} 
        ban_counts = {}  
        
        # Combo Analysis Data Preparation
        hero_partners = {} 
        hero_positions = {} # hero_id -> {pos: count}
        
        for m in matches:
            my_side = 0 if m.is_radiant else 1
            
            # Picks/Bans
            for pb in m.pick_bans:
                if pb.is_pick and pb.team_side == my_side:
                    pick_counts[pb.hero_id] = pick_counts.get(pb.hero_id, 0) + 1
                
                if not pb.is_pick and pb.team_side != my_side:
                    ban_counts[pb.hero_id] = ban_counts.get(pb.hero_id, 0) + 1
            
            # Player Performance for Combos & Positions
            my_team_heroes = []
            my_team_performances = [p for p in m.players if p.team_side == my_side]
            
            for p in my_team_performances:
                my_team_heroes.append(p.hero_id)
                
                # Position Stats
                if p.hero_id not in hero_positions: hero_positions[p.hero_id] = {}
                pos = p.position
                if pos > 0:
                    hero_positions[p.hero_id][pos] = hero_positions[p.hero_id].get(pos, 0) + 1

            # Combo Counting
            for hid in my_team_heroes:
                if hid not in hero_partners: hero_partners[hid] = {}
                for partner in my_team_heroes:
                    if hid != partner:
                        hero_partners[hid][partner] = hero_partners[hid].get(partner, 0) + 1

        # Top Picks / Bans UI
        c_pick, c_ban = st.columns(2)
        
        with c_pick:
            st.subheader("本队常用英雄 (Pick)")
            if pick_counts:
                df_pick = pd.DataFrame(list(pick_counts.items()), columns=['hero_id', 'count'])
                df_pick['英雄'] = df_pick['hero_id'].apply(lambda x: hm.get_hero(x).get('cn_name'))
                df_pick['场次'] = df_pick['count']
                df_pick = df_pick.sort_values('count', ascending=False).head(10)
                st.dataframe(df_pick[['英雄', '场次']], hide_index=True)
            else:
                st.caption("无数据")
                
        with c_ban:
            st.subheader("对手禁用英雄 (Ban)")
            if ban_counts:
                df_ban = pd.DataFrame(list(ban_counts.items()), columns=['hero_id', 'count'])
                df_ban['英雄'] = df_ban['hero_id'].apply(lambda x: hm.get_hero(x).get('cn_name'))
                df_ban['场次'] = df_ban['count']
                df_ban = df_ban.sort_values('count', ascending=False).head(10)
                st.dataframe(df_ban[['英雄', '场次']], hide_index=True)
            else:
                st.caption("无数据")

        # --- Detailed Analysis (Requirement #5 & #6) ---
        st.divider()
        st.subheader("英雄深度分析")
        
        # Hero Selector (sorted by pick count)
        sorted_heroes = sorted(pick_counts.items(), key=lambda x: x[1], reverse=True)
        hero_opts = {f"{hm.get_hero(hid).get('cn_name')} ({count}场)": hid for hid, count in sorted_heroes}
        
        if not hero_opts:
            st.info("暂无英雄数据")
        else:
            sel_hero_label = st.selectbox("选择要分析的英雄", options=list(hero_opts.keys()))
            sel_hero_id = hero_opts[sel_hero_label]
            
            h_data = hm.get_hero(sel_hero_id)
            
            dc1, dc2 = st.columns([1, 3])
            with dc1:
                st.image(h_data.get('img_url'), width=150) # Approx 50% width if column is small
                
                # Position Stats
                st.markdown("**位置分布:**")
                pos_stats = hero_positions.get(sel_hero_id, {})
                if pos_stats:
                    total_p = sum(pos_stats.values())
                    for p_idx in range(1, 6):
                        c = pos_stats.get(p_idx, 0)
                        if c > 0:
                            st.text(f"{p_idx}号位: {c} ({c/total_p*100:.0f}%)")
                else:
                    st.caption("暂无位置数据")

            with dc2:
                st.markdown("**最佳搭档:**")
                partners = hero_partners.get(sel_hero_id, {})
                if partners:
                    # Logic to calculate Win Rate for partners
                    # Need to go back to matches and find when both were picked
                    partner_stats = []
                    
                    for pid, p_count in partners.items():
                        # Count wins where both sel_hero_id and pid were on My Team
                        wins_with_partner = 0
                        for m in matches:
                            my_side = 0 if m.is_radiant else 1
                            # Check if both heroes are in My Picks
                            my_picks = [pb.hero_id for pb in m.pick_bans if pb.is_pick and pb.team_side == my_side]
                            
                            if sel_hero_id in my_picks and pid in my_picks:
                                if m.win:
                                    wins_with_partner += 1
                                    
                        wr = wins_with_partner / p_count if p_count > 0 else 0
                        partner_stats.append({
                            'partner_id': pid,
                            'count': p_count,
                            'win_rate': wr
                        })
                    
                    df_partners = pd.DataFrame(partner_stats)
                    df_partners['搭档'] = df_partners['partner_id'].apply(lambda x: hm.get_hero(x).get('cn_name'))
                    df_partners['头像'] = df_partners['partner_id'].apply(lambda x: hm.get_hero(x).get('icon_url'))
                    df_partners['场次'] = df_partners['count']
                    df_partners['胜率'] = df_partners['win_rate'].apply(lambda x: f"{x:.1%}")
                    
                    df_partners = df_partners.sort_values('count', ascending=False).head(5)
                    
                    st.dataframe(
                        df_partners[['头像', '搭档', '场次', '胜率']],
                        column_config={
                            "头像": st.column_config.ImageColumn("头像", width="small")
                        },
                        hide_index=True
                    )
                else:
                    st.caption("暂无搭档数据")

    # -----------------------------------------------------------------
    # TAB 2: 选手绝活 (With Context Filter - Req #7)
    # -----------------------------------------------------------------
    with tab_player:
        st.subheader("主力选手英雄池")
        
        # Filter Checkbox
        filter_context = st.checkbox("仅分析当前筛选范围内的比赛", value=True)
        
        # Identify Main Players from CURRENT context first
        # Logic Update: Use Manual DB Assignment First
        
        # 1. Pre-calculate auto-detected stats (fallback)
        pos_player_counts = {i: {} for i in range(1, 6)} 
        # Use full matches for detection if context filter is off? 
        # User said "Use recent matches" for conflict resolution.
        # But 'main_players' logic should probably align with the excel export one.
        # Let's use the 'matches' list provided to this view (which is already filtered by time/league).
        
        for m in matches:
            my_side = 0 if m.is_radiant else 1
            my_ps = [p for p in m.players if p.team_side == my_side]
            for p in my_ps:
                if p.position and 1 <= p.position <= 5 and p.account_id:
                    pos_player_counts[p.position][p.account_id] = pos_player_counts[p.position].get(p.account_id, 0) + 1
        
        # 2. Manual DB Lookup
        target_team = db.query(Team).filter(Team.name == selected_team).first()
        manual_players = {i: [] for i in range(1, 6)}
        
        if target_team:
            team_players = db.query(Player).filter(Player.team_id == target_team.team_id, Player.default_pos != None).all()
            for p in team_players:
                if 1 <= p.default_pos <= 5:
                    manual_players[p.default_pos].append(p.account_id)
        
        main_players = {}
        matches_desc = sorted(matches, key=lambda m: m.match_time, reverse=True)
        
        for pos in range(1, 6):
            candidates = manual_players[pos]
            if candidates:
                if len(candidates) == 1:
                    main_players[pos] = candidates[0]
                else:
                    # Conflict: most recent in current matches
                    found = False
                    for m in matches_desc:
                        if found: break
                        my_side = 0 if m.is_radiant else 1
                        my_ps = [p for p in m.players if p.team_side == my_side]
                        for p in my_ps:
                            if p.account_id in candidates:
                                main_players[pos] = p.account_id
                                found = True
                                break
                    if not found:
                        main_players[pos] = candidates[0]
            else:
                # Fallback
                counts = pos_player_counts[pos]
                if counts:
                    main_players[pos] = max(counts, key=counts.get)
        
        pos_tabs = st.tabs([f"{i}号位" for i in range(1, 6)])
        
        for i, tab in enumerate(pos_tabs):
            pos = i + 1
            acc_id = main_players.get(pos)
            
            with tab:
                if not acc_id:
                    st.warning(f"当前范围内未检测到固定的 {pos}号位 选手。")
                    continue
                
                # Get Player Info
                p_info = db.query(Player).filter(Player.account_id == acc_id).first()
                # Or Alias
                alias = db.query(PlayerAlias).filter(PlayerAlias.account_id == acc_id).first()
                p_name_display = f"未知 ({acc_id})"
                if alias and alias.player: p_name_display = alias.player.name
                elif p_info: p_name_display = p_info.name
                
                st.markdown(f"**选手: {p_name_display}**")
                
                # Determine data source
                player_matches = []
                if filter_context:
                    player_matches = matches
                else:
                    # Query ALL matches for this player in last 3 years
                    three_years_ago = datetime.now() - timedelta(days=3*365)
                    perfs = db.query(PlayerPerformance).join(Match).filter(
                        PlayerPerformance.account_id == acc_id,
                        Match.match_time >= three_years_ago
                    ).all()
                    player_matches = [p.match for p in perfs if p.match]
                
                # Calculate Stats
                hero_stats = {} 
                for m in player_matches:
                    p_rec = next((p for p in m.players if p.account_id == acc_id), None)
                    
                    if p_rec:
                        hid = p_rec.hero_id
                        if hid not in hero_stats:
                            hero_stats[hid] = {'picks':0, 'wins':0, 'rad_picks':0, 'rad_wins':0, 'dire_picks':0, 'dire_wins':0}
                        
                        s = hero_stats[hid]
                        s['picks'] += 1
                        
                        radiant_won = (m.is_radiant == m.win)
                        player_won = (p_rec.team_side == 0 and radiant_won) or (p_rec.team_side == 1 and not radiant_won)
                        
                        if player_won: s['wins'] += 1
                        
                        if p_rec.team_side == 0: # Radiant
                            s['rad_picks'] += 1
                            if player_won: s['rad_wins'] += 1
                        else: # Dire
                            s['dire_picks'] += 1
                            if player_won: s['dire_wins'] += 1
                
                if hero_stats:
                    data = []
                    for hid, s in hero_stats.items():
                        total = s['picks']
                        if total == 0: continue
                        
                        wr = s['wins']/total*100
                        rad_wr = (s['rad_wins']/s['rad_picks']*100) if s['rad_picks'] else 0
                        dire_wr = (s['dire_wins']/s['dire_picks']*100) if s['dire_picks'] else 0
                        
                        h = hm.get_hero(hid)
                        data.append({
                            "英雄": h.get('cn_name'),
                            "使用率": f"{(total/len(player_matches)*100):.1f}% ({total})",
                            "胜率": f"{wr:.1f}%",
                            "天辉% (胜率)": f"{(s['rad_picks']/total*100):.0f}% ({rad_wr:.0f}%)",
                            "夜魇% (胜率)": f"{(s['dire_picks']/total*100):.0f}% ({dire_wr:.0f}%)",
                            "icon": h.get('icon_url'),
                            "_sort_pick": total
                        })
                    
                    df = pd.DataFrame(data).sort_values("_sort_pick", ascending=False)
                    
                    st.dataframe(
                        df, 
                        column_config={
                            "icon": st.column_config.ImageColumn("头像", width="small"),
                            "_sort_pick": None 
                        },
                        hide_index=True
                    )
                else:
                    st.info("无数据")

    # -----------------------------------------------------------------
    # TAB 3: BP 链条 (BP Log)
    # -----------------------------------------------------------------
    with tab_bp:
        st.subheader("最近比赛 BP 链条")
        
        # 1. Filter Opponent
        opponents = list(set([m.opponent_name for m in matches]))
        opponents.sort()
        selected_opponents = st.multiselect("过滤对手", options=opponents)
        
        limit_bp = st.number_input("显示最近多少场?", 5, 50, 10)
        
        # Apply filter
        bp_matches = matches
        if selected_opponents:
            bp_matches = [m for m in bp_matches if m.opponent_name in selected_opponents]
            
        bp_matches = bp_matches[:limit_bp]
        
        if not bp_matches:
            st.info("无符合条件的比赛。")
        
        for m in bp_matches:
            res_emoji = "✅" if m.win else "❌"
            # Determine side for display
            my_side_str = "天辉" if m.is_radiant else "夜魇"
            header = f"{m.match_time.strftime('%m-%d')} | vs {m.opponent_name} ({my_side_str}) | {res_emoji}"
            
            with st.expander(header, expanded=True):
                rad_name = m.team_name if m.is_radiant else m.opponent_name
                dire_name = m.opponent_name if m.is_radiant else m.team_name
                
                is_radiant_first = (m.is_radiant == m.first_pick)
                
                # Center the visual - or standard layout
                # User requested "side-by-side" compact view for BP chain too?
                # "优化一下比赛列表和统计分析中BP显示的排版" -> "Stats Analysis BP Chain" also implied.
                # So we use the new layout here too.
                
                render_bp_visual(m.pick_bans, rad_name, dire_name, hm, first_pick_radiant=is_radiant_first, layout="side-by-side")

    db.close()